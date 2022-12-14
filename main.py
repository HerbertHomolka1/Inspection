from openpyxl import Workbook, load_workbook
import kivy
import os


from kivy.uix.behaviors import ToggleButtonBehavior

from kivy.uix.screenmanager import ScreenManager, Screen

from kivymd.uix.boxlayout import MDBoxLayout, BoxLayout

from kivymd.uix.scrollview import ScrollView

from kivy.app import App
from kivymd.uix.behaviors.toggle_behavior import MDToggleButton
from kivymd.uix.button import MDFlatButton, MDRectangleFlatButton



from kivy.lang import Builder

from kivymd.app import MDApp



class MyToggleButton(MDFlatButton, MDToggleButton):
    def __init__(self, **kwargs):
        super(MyToggleButton, self).__init__(**kwargs)



class MainScreen(Screen):


    operator = kivy.properties.StringProperty('')
    cell_number = kivy.properties.StringProperty('')
    shop_order = kivy.properties.StringProperty('')

    def dirdir(self):
        t = str(os.getcwd())
        for x in list(self.children)[0].children:
            if isinstance(x,MDBoxLayout):
                list(x.children)[0].text=t


    def next_screen(self, shop_order, cell_number, operator):

        check1 = operator in ['HHO', 'WLE', 'TPE']
        check2 = len(str(shop_order)) == 5 and shop_order.isnumeric()
        try:
            check3 =  int(cell_number) < 91 and int(cell_number) > 14

        except ValueError:
            check3 = False

        if check1 and check2  and check3:
            global c_number
            global s_order

            c_number = int(cell_number)
            s_order = shop_order

            self.manager.current = 'Screen2'
            self.create_new_excell(shop_order, cell_number, operator)




    def create_new_excell(self,shop_order, cell_number, operator ):

        wb = Workbook()
        ws = wb.active
        ws.title = 'cell stack'
        ws.append(['shop order','number of cells', 'operator'])
        ws.append([shop_order, cell_number, operator])

        ws.append([str(i) for i in range(1,24)])


        wb.save('example.xlsx')



class MyButton(ToggleButtonBehavior, MDRectangleFlatButton):

    def __init__(self, **kwargs):
        super(MyButton, self).__init__(**kwargs)


    def on_state(self, widget, value):
        if value == 'down':
            self.md_bg_color= ( 0.5, 0.4, 0.2)
        else:
            self.md_bg_color= (0,0,0,1)

class Screen3(Screen):
    def close_app_now(self):
        App.get_running_app().stop()

# class Screen4(Screen):
#     pass

# time.sleep(1)
# App.get_running_app().stop()

class Screen2(Screen):

    cell_counter= kivy.properties.NumericProperty(1)
    def go_to_previous_cell(self):
        if self.cell_counter >1:
            self.cell_counter -=1
            a = self.get_grid_btns()
            #/storage/sdcard0/Android/data/package/files/Download/
            wb = load_workbook('example.xlsx')
            ws = wb.active
            states = [ws[chr(65+x)+str(self.cell_counter+3)].value for x in range(len(a)) ]
            states1 = ['normal' if x== 'ok' else 'down' for x in states]
            for i,btn in enumerate(a):
                btn.state = states1[i]

    def insert_excell_row(self,lista):
        wb = load_workbook('example.xlsx')
        ws = wb.active
        ws.append(['!!!!!!!' if x.state == 'down' else 'ok' for x in lista])
        wb.save('example.xlsx')

    def go_to_next_cell(self):
        global c_number
        print(c_number)

        wb = load_workbook('example.xlsx')
        ws = wb.active
        if ws['A'+str(self.cell_counter+1+3)].value:
            self.cell_counter +=1
            a = self.get_grid_btns()
            states = [ws[chr(65+x)+str(self.cell_counter+3)].value for x in range(len(a)) ]
            states1 = ['normal' if x== 'ok' else 'down' for x in states]
            for i,btn in enumerate(a):
                btn.state = states1[i]

        else:
            self.cell_counter +=1
            a = self.get_grid_btns()
            self.insert_excell_row(self.get_grid_btns())
            self.set_to_normal(a)

        if self.cell_counter > c_number:
            global s_order
            wb.save('example.xlsx')
            #wb.save('/storage/emulated/0/Download/'+s_order+'.xlsx')
            self.manager.current = 'Screen3'


                # wb = load_workbook('example.xlsx')
        # ws = wb.active
        # self.cell_counter += 1
        # if ws['A'+str(self.cell_counter+3)].value:
        #     pass
        #     # a = self.get_grid_btns()
        #     # states = [ws[chr(65+x)+str(self.cell_counter+3)].value for x in range(len(a)) ]
        #     # states1 = ['normal' if x== 'ok' else 'down' for x in states]
        #     # for i,btn in enumerate(a):
        #     #     btn.state = states1[i]
        #
        # else:
        #
        #     a = self.get_grid_btns()
        #     self.set_to_normal(a)

    def set_to_normal(self,btns):
        for btn in btns:
            btn.state = 'normal'
    def get_grid_btns(self):
        a = []
        for child in self.children:
            for c in child.children:

                if isinstance(c, ScrollView):
                    for kid in c.children:

                        for btn in  kid.children:

                            a.append(btn)
        return a



sm = ScreenManager()
sm.add_widget(MainScreen(name = 'MainScreen'))
sm.add_widget((Screen2(name= 'Screen2')))

class HHoCumminsApp(MDApp):
    def build(self):
        #return Screen4
        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "Orange"
        screen = Builder.load_file('HHO.kv')
        return screen

HHoCumminsApp().run()
